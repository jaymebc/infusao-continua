#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel to database.js Converter
Converte Drug-Infusion-Database Excel para database.js JavaScript
"""

import openpyxl
import json
import sys
from datetime import datetime

def parse_boolean(value):
    """Converte string TRUE/FALSE para boolean JavaScript"""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.upper() == "TRUE"
    return bool(value)

def parse_number(value):
    """Converte para número ou None"""
    if value is None or value == "":
        return None
    try:
        # Tenta float primeiro
        num = float(value)
        # Se for inteiro, retorna como int
        if num.is_integer():
            return int(num)
        return num
    except (ValueError, TypeError):
        return None

def parse_json_field(value):
    """Parseia campo JSON da planilha"""
    if value is None or value == "":
        return None
    
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError as e:
            print(f"⚠️  AVISO: JSON inválido: {value[:50]}... - Erro: {e}")
            return None
    
    return value

def sanitize_string(value):
    """Limpa e sanitiza strings"""
    if value is None:
        return ""
    return str(value).strip()

def generate_js_value(value, indent=0):
    """Converte valor Python para string JavaScript formatada"""
    indent_str = "    " * indent
    
    if value is None:
        return "null"
    elif isinstance(value, bool):
        return "true" if value else "false"
    elif isinstance(value, (int, float)):
        return str(value)
    elif isinstance(value, str):
        # Escapa aspas e quebras de linha
        escaped = value.replace(chr(92), chr(92)+chr(92)).replace('"', chr(92)+'"').replace(chr(10), chr(92)+'n')
        return f'"{escaped}"'
    elif isinstance(value, list):
        if not value:
            return "[]"
        # Array de objetos (presentation_options)
        items = []
        for item in value:
            items.append(generate_js_value(item, indent + 1))
        return "[\n" + indent_str + "    " + (",\n" + indent_str + "    ").join(items) + "\n" + indent_str + "]"
    elif isinstance(value, dict):
        if not value:
            return "{}"
        items = []
        for key, val in value.items():
            js_val = generate_js_value(val, indent + 1)
            items.append(f'{key}: {js_val}')
        return "{ " + ", ".join(items) + " }"
    else:
        return f'"{str(value)}"'

def excel_to_database(input_file, output_file):
    """Converte Excel para database.js"""
    
    print("=" * 70)
    print("📊 CONVERSOR: EXCEL → DATABASE.JS")
    print("=" * 70)
    print(f"📂 Lendo arquivo: {input_file}")
    
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        
        # Ler cabeçalhos (primeira linha)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"✅ Colunas encontradas: {len(headers)}")
        print(f"📋 Cabeçalhos: {', '.join(headers[:5])}...")
        
        # Dicionário para armazenar as drogas
        drugs = {}
        
        # Processar linhas (a partir da linha 2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Se drug_key está vazio, pular linha
                continue
            
            # Criar dicionário com os dados da linha
            drug_data = {}
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                if header == "has_presentation_selector":
                    drug_data[header] = parse_boolean(value)
                elif header == "presentation_options_json":
                    drug_data[header] = parse_json_field(value)
                elif header in ["default_quantity", "default_volume", "dose_min", "dose_max"]:
                    drug_data[header] = parse_number(value)
                else:
                    drug_data[header] = sanitize_string(value) if value else ""
            
            drug_key = drug_data["drug_key"]
            drugs[drug_key] = drug_data
            
            print(f"  ✓ Linha {row_idx}: {drug_key}")
        
        print(f"\n✅ Total de drogas processadas: {len(drugs)}")
        
        # Gerar código JavaScript
        print(f"\n📝 Gerando {output_file}...")
        
        js_code = generate_javascript(drugs)
        
        # Salvar arquivo
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(js_code)
        
        print(f"✅ Arquivo salvo: {output_file}")
        print(f"📏 Tamanho: {len(js_code)} caracteres")
        print("\n🎉 CONVERSÃO CONCLUÍDA COM SUCESSO!")
        
    except FileNotFoundError:
        print(f"❌ ERRO: Arquivo '{input_file}' não encontrado!")
        sys.exit(1)
    except Exception as e:
        print(f"❌ ERRO: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def generate_javascript(drugs):
    """Gera código JavaScript do database.js"""
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    js = f"""// database.js - Drug Infusion Database
// Gerado automaticamente em: {timestamp}
// NÃO EDITE MANUALMENTE - Use o script excel_to_database.py

export const rawDrugDatabase = {{
"""
    
    for drug_key, drug_data in drugs.items():
        js += f"    {drug_key}: {{\n"
        
        # Campos obrigatórios primeiro
        js += f'        drug_key: "{drug_data["drug_key"]}",\n'
        js += f'        group_key: "{drug_data["group_key"]}",\n'
        js += f'        calc_type: "{drug_data["calc_type"]}",\n'
        js += f'        name: "{drug_data["name"]}",\n'
        
        # brand_name (pode ser vazio)
        if drug_data.get("brand_name"):
            js += f'        brand_name: "{drug_data["brand_name"]}",\n'
        
        js += f'        category: "{drug_data["category"]}",\n'
        js += f'        presentation: "{drug_data["presentation"]}",\n'
        js += f'        dose_unit: "{drug_data["dose_unit"]}",\n'
        js += f'        default_quantity_unit: "{drug_data["default_quantity_unit"]}",\n'
        js += f'        default_quantity: {drug_data["default_quantity"]},\n'
        js += f'        default_volume: {drug_data["default_volume"]},\n'
        js += f'        dose_range_text: "{drug_data["dose_range_text"]}",\n'
        
        # ✅ ADICIONAR dose_min e dose_max
        if drug_data.get("dose_min") is not None:
            js += f'        dose_min: {drug_data["dose_min"]},\n'
        
        if drug_data.get("dose_max") is not None:
            js += f'        dose_max: {drug_data["dose_max"]},\n'
        
        # Campos opcionais
        if drug_data.get("bolus_type"):
            js += f'        bolus_type: "{drug_data["bolus_type"]}",\n'
        
        # has_presentation_selector
        has_pres = drug_data.get("has_presentation_selector")
        if has_pres is not None:
            js += f'        has_presentation_selector: {generate_js_value(has_pres)},\n'
        
        # presentation_options (do JSON)
        pres_options = drug_data.get("presentation_options_json")
        if pres_options:
            js += f'        presentation_options: {generate_js_value(pres_options, indent=2)},\n'
        
        # notes
        if drug_data.get("notes"):
            js += f'        notes: "{drug_data["notes"]}"\n'
        else:
            # Remove vírgula do último campo
            js = js.rstrip(",\n") + "\n"
        
        js += "    },\n"
    
    js += "};\n"
    
    return js

if __name__ == "__main__":
    input_file = "Drug-Infusion-Database-v2.1-FIXED.xlsx"
    output_file = "database.js"
    
    excel_to_database(input_file, output_file)
    
    print("\n" + "=" * 70)
    print("📤 PRÓXIMOS PASSOS:")
    print("=" * 70)
    print("1. ✅ Substitua o arquivo database.js no seu projeto")
    print("2. 🧪 Teste o app para verificar se tudo funciona")
    print("3. 🔄 Sempre que atualizar o Excel, rode este script novamente")
    print("=" * 70)